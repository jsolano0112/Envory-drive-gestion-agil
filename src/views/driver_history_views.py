# views/driver_history_views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from datetime import datetime, timedelta
import csv
import io
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..utils.decorators import admin_required, cliente_required, conductor_required, get_user_type

from ..models.models import Conductor

@login_required
def detalle_conductor(request, id):
    conductor = get_object_or_404(Conductor, pk=id)

    # Preferir placa si existe, sino documento, sino nombre completo
    if hasattr(conductor, 'vehiculo') and conductor.vehiculo and getattr(conductor.vehiculo, 'placa', ''):
        search_value = conductor.vehiculo.placa
    else:
        search_value = conductor.numero_documento or f"{conductor.user.first_name} {conductor.user.last_name}"

    url = reverse('driver_history')   # debe ser name='driver_history' en urls.py
    return redirect(f"{url}?search={search_value}")

@admin_required
@login_required
def driver_history(request):
    """Vista principal del historial de conductor"""
    context = {
        'driver': None,
        'search_term': '',
        'active_tab': 'viajes',
        'date_start': '',
        'date_end': ''
    }
    
    search_term = request.GET.get('search', '')
    active_tab = request.GET.get('tab', 'viajes')
    date_start = request.GET.get('date_start', '')
    date_end = request.GET.get('date_end', '')
    
    if search_term:
        # Buscar conductor por nombre, documento o placa
        driver = Conductor.objects.filter(
            Q(user__first_name__icontains=search_term) |
            Q(user__last_name__icontains=search_term) |
            Q(numero_documento__icontains=search_term) |
            Q(vehiculo__placa__icontains=search_term)
        ).select_related('user', 'vehiculo').first()
        
        if driver:
            # ========== GENERAR DATA MOCKEADA DE VIAJES ==========
            trips = []
            base_date = datetime.now()

            for i in range(5):  # Cambia este número para más o menos viajes
                # Fechas aleatorias en los últimos 90 días
                days_ago = random.randint(0, 90)
                trip_date = base_date - timedelta(days=days_ago, hours=random.randint(6, 22))
                
                # Estados posibles con probabilidades realistas
                status_weights = [('Completado', 0.85), ('Cancelado', 0.10), ('En Progreso', 0.04), ('Aceptado', 0.01)]
                status = random.choices([s[0] for s in status_weights], weights=[s[1] for s in status_weights])[0]
                
                # Calificación solo para viajes completados
                rating = random.choice([4.5, 5.0, 4.0, 4.5, 3.5, 5.0, 4.0, 4.5, 5.0, 3.0]) if status == 'Completado' and random.random() > 0.1 else None
                
                # Ubicaciones realistas en Bogotá
                locations = [
                    'Centro Internacional', 'Aeropuerto El Dorado', 'Zona Rosa', 'Parque de la 93',
                    'Centro Andino', 'Gran Estación', 'Plaza de las Américas', 'Calle 80',
                    'Centro Empresarial Santa Fe', 'Parque Nacional', 'Usaquén', 'Chapinero',
                    'Salitre Plaza', 'Portal Norte', 'Calle 170', 'Centro Histórico'
                ]
                
                origin = random.choice(locations)
                destination = random.choice([loc for loc in locations if loc != origin])
                
                # Clientes mockeados
                clients = [
                    'María González', 'Carlos Rodríguez', 'Ana López', 'Juan Martínez', 
                    'Laura Sánchez', 'Diego Pérez', 'Camila Torres', 'Andrés Ramírez',
                    'Valentina Morales', 'Santiago Herrera', 'Isabella Castro', 'Mateo Vargas'
                ]
                
                trip = {
                    'id': f'TRP{1000 + i}',
                    'date': trip_date,
                    'status': status,
                    'rating': rating,
                    'origin': origin,
                    'destination': destination,
                    'client': random.choice(clients),
                    'amount': random.randint(15000, 45000),
                    'distance': round(random.uniform(5.0, 25.0), 1),
                    'duration': f"{random.randint(20, 90):02d}:{random.randint(0, 59):02d}:00"
                }
                trips.append(trip)

            # ========== OBTENER HISTORIAL DE ESTADOS ==========
            try:
                from ..models.models import HistorialEstadoConductor
                status_history = []
                
                # Obtener los últimos 20 cambios de estado
                historial_estados = HistorialEstadoConductor.objects.filter(
                    conductor=driver
                ).select_related('usuario_modificador').order_by('-fecha_cambio')[:20]
                
                for cambio in historial_estados:
                    # Crear títulos descriptivos según los estados
                    titulos = {
                        ('Pendiente', 'En Corrección'): 'Documentos Pendientes de Corrección',
                        ('En Corrección', 'Pendiente'): 'Correcciones Enviadas',
                        ('Pendiente', 'Activo'): 'Conductor Aprobado y Activado',
                        ('Pendiente', 'Rechazado'): 'Conductor Rechazado',
                        ('Activo', 'Inactivo'): 'Conductor Desactivado',
                        ('Inactivo', 'Activo'): 'Conductor Reactivado',
                        ('Activo', 'Suspendido'): 'Conductor Suspendido',
                        ('Suspendido', 'Activo'): 'Suspensión Levantada',
                        ('Activo', 'Bloqueado'): 'Conductor Bloqueado',
                        ('Bloqueado', 'Activo'): 'Conductor Desbloqueado',
                        ('Activo', 'Dado de Baja'): 'Conductor Dado de Baja',
                    }
                    
                    title = titulos.get(
                        (cambio.estado_anterior, cambio.estado_nuevo),
                        f'Cambio de Estado: {cambio.estado_anterior} → {cambio.estado_nuevo}'
                    )
                    
                    status_history.append({
                        'id': cambio.id,
                        'date': cambio.fecha_cambio,
                        'title': title,
                        'previous_status': cambio.estado_anterior,
                        'new_status': cambio.estado_nuevo,
                        'reason': cambio.descripcion_motivo or cambio.get_motivo_display(),
                        'modified_by': cambio.usuario_modificador.get_full_name() if cambio.usuario_modificador else 'Sistema'
                    })
            
            except (ImportError, AttributeError):
                # Si el modelo no existe aún, usar data mockeada como fallback
                status_history = []
                status_sequence = ['Pendiente', 'Activo', 'Suspendido', 'Activo', 'Inactivo', 'Activo']

                current_date = driver.fecha_registro
                admins = ['Admin Sistema', 'María González', 'Carlos Rodríguez', 'Supervisor General']

                for i, new_status in enumerate(status_sequence):
                    if i == 0:
                        prev_status = 'Pendiente'
                        title = 'Registro Inicial'
                        reason = 'Registro inicial del conductor en la plataforma'
                    else:
                        prev_status = status_sequence[i-1]
                        transitions = {
                            ('Pendiente', 'Activo'): ('Conductor Aprobado', 'Documentos verificados correctamente'),
                            ('Activo', 'Suspendido'): ('Conductor Suspendido', 'Queja de cliente por demora excesiva'),
                            ('Suspendido', 'Activo'): ('Suspensión Levantada', 'Cumplimiento de condiciones de reactivación'),
                            ('Activo', 'Inactivo'): ('Conductor Desactivado', 'Por decisión propia'),
                            ('Inactivo', 'Activo'): ('Conductor Reactivado', 'Solicitud de reactivación aprobada'),
                        }
                        title, reason = transitions.get((prev_status, new_status), ('Cambio de Estado', 'Cambio administrativo'))
                    
                    days_later = random.randint(5, 30) if i > 0 else 0
                    current_date = current_date + timedelta(days=days_later)
                    
                    change = {
                        'id': i + 1,
                        'date': current_date,
                        'title': title,
                        'previous_status': prev_status,
                        'new_status': new_status,
                        'reason': reason,
                        'modified_by': random.choice(admins)
                    }
                    status_history.append(change)

            # ========== CALCULAR ESTADÍSTICAS ==========
            total_trips = len(trips)
            completed_trips = len([t for t in trips if t['status'] == 'Completado'])
            rated_trips = [t for t in trips if t['rating']]
            avg_rating = sum(t['rating'] for t in rated_trips) / len(rated_trips) if rated_trips else 0
            completion_rate = int((completed_trips / total_trips * 100)) if total_trips > 0 else 0

            stats = {
                'total_trips': total_trips,
                'avg_rating': round(avg_rating, 1) if avg_rating > 0 else 4.8,
                'completion_rate': completion_rate,
                'total_reports': random.randint(0, 5)
            }
            
            # ========== ACTUALIZAR CONTEXTO ==========
            context.update({
                'driver': driver,
                'stats': stats,
                'trips': trips,
                'status_history': status_history,
                'search_term': search_term,
                'active_tab': active_tab,
                'date_start': date_start,
                'date_end': date_end
            })
    
    return render(request, 'conductores/driver_details.html', context)

@admin_required
@csrf_exempt
@require_http_methods(["POST"])
def update_driver_status(request, driver_id):
    try:
        driver = get_object_or_404(Conductor, id=driver_id)
        new_status = request.POST.get('status')
        
        # Validar estado
        valid_statuses = ['Activo', 'Inactivo', 'Suspendido', 'Bloqueado', 'En Revisión']
        if new_status not in valid_statuses:
            return JsonResponse({
                'success': False,
                'message': 'Estado no válido'
            }, status=400)
        
        # Guardar estado anterior
        old_status = driver.estado
        
        # Actualizar estado del conductor
        driver.estado = new_status
        driver.save()
        
        # Registrar el cambio en el historial
        try:
            from ..models.models import HistorialEstadoConductor
            
            HistorialEstadoConductor.objects.create(
                conductor=driver,
                estado_anterior=old_status,
                estado_nuevo=new_status,
                motivo='Manual',
                descripcion_motivo=request.POST.get('reason', 'Cambio manual por administrador'),
                usuario_modificador=request.user
            )
        except (ImportError, AttributeError):
            # Si el modelo no existe aún, continuar sin registrar
            pass
        
        return redirect(f"{request.META.get('HTTP_REFERER', '/')}?search={driver.numero_documento}")
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al actualizar estado: {str(e)}'
        }, status=500)

@admin_required
@csrf_exempt
@require_http_methods(["POST"])
def generate_report(request, driver_id):
    try:
        driver = get_object_or_404(Conductor, id=driver_id)
        report_type = request.POST.get('report_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Validar campos requeridos
        if not report_type or not start_date or not end_date:
            return JsonResponse({
                'success': False,
                'message': 'Faltan campos requeridos'
            }, status=400)
        
        # Convertir fechas
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Validar que la fecha de inicio sea anterior a la fecha final
        if start_date_obj > end_date_obj:
            return JsonResponse({
                'success': False,
                'message': 'La fecha de inicio debe ser anterior a la fecha final'
            }, status=400)
        
        # Crear workbook de Excel
        wb = Workbook()
        ws = wb.active
        
        # Estilos para el encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Información del reporte (primera fila)
        ws['A1'] = f"Reporte {report_type.title()}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:K1')
        
        # Información del conductor
        ws['A3'] = "Información del Conductor"
        ws['A3'].font = Font(bold=True)
        ws.merge_cells('A3:K3')
        
        conductor_info = [
            ["Nombre:", driver.get_nombre_completo()],
            ["Documento:", driver.numero_documento],
            ["Placa:", driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else 'N/A'],
            ["Período:", f"{start_date} al {end_date}"],
        ]
        
        for i, (label, value) in enumerate(conductor_info, 5):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = value
            ws.merge_cells(f'B{i}:K{i}')
        
        # Espacio en blanco
        current_row = len(conductor_info) + 6
        
        if report_type in ['monthly', 'daily']:
            # Reporte de viajes
            ws[f'A{current_row}'] = "REPORTE DE VIAJES"
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            ws.merge_cells(f'A{current_row}:K{current_row}')
            
            # Encabezados
            headers = [
                "ID Conductor", "Nombre Conductor", "Placa Vehículo", "ID Cliente", 
                "Nombre Cliente", "Lugar Recogida", "Destino", "Distancia (km)", 
                "Tiempo Acumulado", "Fecha Inicio", "Fecha Fin"
            ]
            
            header_row = current_row + 2
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = border
            
            # Datos de ejemplo
            sample_trips = [
                [driver.id, driver.get_nombre_completo(), driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else 'N/A', 
                 "CLI001", "Juan Pérez", "Centro", "Aeropuerto", 15.5, "02:30:00", "2025-01-15 08:00", "2025-01-15 10:30"],
                [driver.id, driver.get_nombre_completo(), driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else 'N/A',
                 "CLI002", "María González", "Mall", "Hospital", 8.2, "01:45:00", "2025-01-16 14:00", "2025-01-16 15:45"],
                [driver.id, driver.get_nombre_completo(), driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else 'N/A',
                 "CLI003", "Carlos Rodríguez", "Estadio", "Centro", 12.8, "02:15:00", "2025-01-17 19:00", "2025-01-17 21:15"],
            ]
            
            # Agregar datos
            for row_idx, trip in enumerate(sample_trips, header_row + 1):
                for col_idx, value in enumerate(trip, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
            
            # Total de viajes
            total_row = header_row + len(sample_trips) + 2
            ws[f'A{total_row}'] = f"TOTAL VIAJES: {len(sample_trips)}"
            ws[f'A{total_row}'].font = Font(bold=True)
            ws.merge_cells(f'A{total_row}:K{total_row}')
            
        elif report_type == 'monetary':
            # Reporte monetario
            ws[f'A{current_row}'] = "REPORTE MONETARIO"
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            ws.merge_cells(f'A{current_row}:K{current_row}')
            
            # Encabezados
            headers = ["ID Conductor", "Nombre Conductor", "Fecha Inicio", "Fecha Fin", "Valor Total Generado"]
            
            header_row = current_row + 2
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = border
            
            # Datos de ejemplo
            total_value = 2850000
            
            data_row = header_row + 1
            ws.cell(row=data_row, column=1).value = driver.id
            ws.cell(row=data_row, column=2).value = driver.get_nombre_completo()
            ws.cell(row=data_row, column=3).value = start_date
            ws.cell(row=data_row, column=4).value = end_date
            ws.cell(row=data_row, column=5).value = f"${total_value:,.0f}"
            
            # Aplicar bordes
            for col in range(1, 6):
                ws.cell(row=data_row, column=col).border = border
        
        # Ajustar ancho de columnas
        for col in range(1, 12):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Crear respuesta HTTP
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'reporte_{report_type}_{driver.numero_documento}_{start_date}_{end_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al generar reporte: {str(e)}'
        }, status=500)

@admin_required
def export_history(request, driver_id):
    """Exporta el historial completo del conductor a CSV"""
    try:
        driver = get_object_or_404(Conductor, id=driver_id)
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="historial_{driver.numero_documento}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        writer.writerow(['INFORMACIÓN DEL CONDUCTOR'])
        writer.writerow(['Nombre Completo', driver.get_nombre_completo()])
        writer.writerow(['Documento', driver.numero_documento])
        writer.writerow(['Email', driver.user.email])
        writer.writerow(['Estado', driver.estado])
        writer.writerow(['Fecha Registro', driver.fecha_registro.strftime('%Y-%m-%d')])
        writer.writerow([])
        
        if hasattr(driver, 'vehiculo') and driver.vehiculo:
            writer.writerow(['INFORMACIÓN DEL VEHÍCULO'])
            writer.writerow(['Placa', driver.vehiculo.placa])
            writer.writerow(['Marca', driver.vehiculo.marca])
            writer.writerow(['Modelo', driver.vehiculo.modelo])
            writer.writerow([])
        
        return response
    
    except Exception as e:
        return HttpResponse(f'Error al exportar: {str(e)}', status=500)

@admin_required
@require_http_methods(["GET"])
def driver_statistics_api(request, driver_id):
    """Endpoint API para obtener estadísticas del conductor"""
    try:
        driver = get_object_or_404(Conductor, id=driver_id)
        
        statistics = {
            'conductor': {
                'id': driver.id,
                'nombre_completo': driver.get_nombre_completo(),
                'numero_documento': driver.numero_documento,
                'email': driver.user.email,
                'estado': driver.estado,
                'fecha_registro': driver.fecha_registro.strftime('%Y-%m-%d')
            },
            'vehiculo': {
                'placa': driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else None,
                'marca': driver.vehiculo.marca if hasattr(driver, 'vehiculo') and driver.vehiculo else None,
                'modelo': driver.vehiculo.modelo if hasattr(driver, 'vehiculo') and driver.vehiculo else None,
            } if hasattr(driver, 'vehiculo') and driver.vehiculo else None,
            'stats': {
                'total_viajes': 247,
                'calificacion_promedio': 4.8,
                'tasa_completado': 96,
                'total_reportes': 2,
            }
        }
        
        return JsonResponse({'success': True, 'data': statistics})
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@admin_required
@require_http_methods(["GET"])
def driver_autocomplete_api(request):
    """API para autocompletado de conductores"""
    try:
        search_term = request.GET.get('q', '').strip()
        
        if not search_term or len(search_term) < 2:
            return JsonResponse({'results': []})
        
        conductores = Conductor.objects.select_related('user', 'vehiculo').filter(
            Q(user__first_name__icontains=search_term) |
            Q(user__last_name__icontains=search_term) |
            Q(segundo_nombre__icontains=search_term) |
            Q(segundo_apellido__icontains=search_term) |
            Q(numero_documento__icontains=search_term) |
            Q(vehiculo__placa__icontains=search_term)
        )[:10]
        
        results = []
        for conductor in conductores:
            placa = conductor.vehiculo.placa if hasattr(conductor, 'vehiculo') and conductor.vehiculo else 'Sin placa'
            results.append({
                'id': conductor.id,
                'nombre_completo': conductor.get_nombre_completo(),
                'numero_documento': conductor.numero_documento,
                'placa': placa,
                'estado': conductor.estado
            })
        
        return JsonResponse({'results': results})
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)