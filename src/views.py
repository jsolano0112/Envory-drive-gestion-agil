from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import transaction, IntegrityError
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import json
import re
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout

from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect, get_object_or_404

from .models import Cliente, Compania, Conductor, Vehiculo, DocumentoConductor, Viaje, Novedad
from .decorators import admin_required, cliente_required, conductor_required, get_user_type


@admin_required
@login_required
def conductores_todos(request):
    """
    Lista todos los conductores (sin filtros) y renderiza la plantilla.
    """
    # Ajusta select_related/prefetch_related según tus relaciones reales
    conductores = Conductor.objects.select_related('user').all()
    return render(request, 'conductores/listado_todos.html', {'conductores': conductores})

@admin_required
@login_required
def detalle_conductor(request, id):
    """
    Detalle simple de conductor. Mantener si lo necesitas en otras partes.
    """
    conductor = get_object_or_404(Conductor, id=id)
    viajes = Viaje.objects.filter(conductor=conductor)
    return render(request, 'conductores/detalle.html', {'conductor': conductor, 'viajes': viajes})




def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, "¡Bienvenido a EVORY DRIVE!")
            return redirect('home')
        else:
            messages.error(request, "Usuario o contraseña incorrectos.")
            return redirect('login')

    return render(request, 'login.html')


# ====================================
# VISTAS EXISTENTES
# ====================================
@login_required
def home_view(request):
    user_type = get_user_type(request.user)
    return render(request, 'home.html', {'user_type': user_type})

def logout_view(request):
    logout(request)
    return redirect('inicio') 

def index(request):
    return render(request, 'index.html')


def inicio(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')  # redirige a la vista protegida
        else:
            messages.error(request, 'Credenciales inválidas. Intenta nuevamente.')
            return render(request, 'login.html')
    
    return render(request, 'login.html')

@login_required
def driver_registration(request):
    return render(request, 'driver_registration.html')



# ====================================
# MÓDULO: REGISTRO DE CLIENTES
# ====================================

@admin_required
@login_required
def client_registration(request):
    """
    Vista que renderiza el formulario de registro de clientes.
    GET: Muestra el formulario
    """
    return render(request, 'client_registration.html')


@admin_required
@csrf_exempt
@require_http_methods(["POST"])
def client_registration_api(request):
    """
    Endpoint POST para registrar nuevos clientes.
    Valida datos, verifica duplicados y crea el usuario + perfil de cliente.
    
    Respuesta JSON:
    - success: true/false
    - message: mensaje descriptivo
    - data: información adicional (opcional)
    """
    try:
        # Parsear datos JSON del request
        data = json.loads(request.body)
        
        # ===================================
        # VALIDACIÓN DE CAMPOS OBLIGATORIOS
        # ===================================
        required_fields = [
            'primer_nombre', 'primer_apellido', 'tipo_documento',
            'numero_documento', 'correo', 'telefono', 'compania_id', 'password'
        ]
        
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return JsonResponse({
                'success': False,
                'message': 'Error: faltan campos obligatorios',
                'missing_fields': missing_fields
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE FORMATO DE NOMBRES
        # ===================================
        # Solo letras y espacios (sin números ni símbolos)
        name_pattern = re.compile(r'^[A-Za-zÁÉÍÓÚáéíóúÑñ\s]+$')
        
        if not name_pattern.match(data['primer_nombre']):
            return JsonResponse({
                'success': False,
                'message': 'El primer nombre no debe contener números ni símbolos'
            }, status=400)
        
        if not name_pattern.match(data['primer_apellido']):
            return JsonResponse({
                'success': False,
                'message': 'El primer apellido no debe contener números ni símbolos'
            }, status=400)
        
        if data.get('segundo_nombre') and not name_pattern.match(data['segundo_nombre']):
            return JsonResponse({
                'success': False,
                'message': 'El segundo nombre no debe contener números ni símbolos'
            }, status=400)
        
        if data.get('segundo_apellido') and not name_pattern.match(data['segundo_apellido']):
            return JsonResponse({
                'success': False,
                'message': 'El segundo apellido no debe contener números ni símbolos'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE EMAIL
        # ===================================
        email = data['correo'].lower().strip()
        try:
            validate_email(email)
        except ValidationError:
            return JsonResponse({
                'success': False,
                'message': 'Verificar correo'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE CONTRASEÑA
        # ===================================
        password = data['password']
        confirm_password = data.get('confirm_password', '')
        
        # Verificar coincidencia
        if password != confirm_password:
            return JsonResponse({
                'success': False,
                'message': 'Las contraseñas no coinciden'
            }, status=400)
        
        # Validar requisitos: mínimo 8 caracteres, mayúscula, minúscula, número
        if len(password) < 8:
            return JsonResponse({
                'success': False,
                'message': 'La contraseña debe tener mínimo 8 caracteres'
            }, status=400)
        
        if not re.search(r'[A-Z]', password):
            return JsonResponse({
                'success': False,
                'message': 'La contraseña debe contener al menos una mayúscula'
            }, status=400)
        
        if not re.search(r'[a-z]', password):
            return JsonResponse({
                'success': False,
                'message': 'La contraseña debe contener al menos una minúscula'
            }, status=400)
        
        if not re.search(r'\d', password):
            return JsonResponse({
                'success': False,
                'message': 'La contraseña debe contener al menos un número'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE NÚMERO DE DOCUMENTO
        # ===================================
        numero_documento = data['numero_documento'].strip()
        if not numero_documento.isdigit():
            return JsonResponse({
                'success': False,
                'message': 'El número de documento solo debe contener números'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE TELÉFONO
        # ===================================
        telefono = data['telefono'].strip()
        if not re.match(r'^[0-9]{10}$', telefono):
            return JsonResponse({
                'success': False,
                'message': 'El teléfono debe tener 10 dígitos'
            }, status=400)
        
        # ===================================
        # VERIFICAR USUARIO NO EXISTENTE
        # ===================================
        
        # Verificar por email
        if User.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'message': 'El email se encuentra registrado'
            }, status=409)
        
        # Verificar por número de documento
        if Cliente.objects.filter(numero_documento=numero_documento).exists():
            return JsonResponse({
                'success': False,
                'message': 'El numero de documento se encuentra registrado'
            }, status=409)
        
        # ===================================
        # VERIFICAR QUE LA COMPAÑÍA EXISTE
        # ===================================
        try:
            compania = Compania.objects.get(id=data['compania_id'], estado=True)
        except Compania.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'La compañía seleccionada no es válida'
            }, status=400)
        
        # ===================================
        # CREAR USUARIO Y CLIENTE
        # ===================================
        with transaction.atomic():
            # Crear usuario de Django
            user = User.objects.create_user(
                username=email,  # Usamos el email como username
                email=email,
                password=password,
                first_name=data['primer_nombre'].strip(),
                last_name=data['primer_apellido'].strip()
            )
            
            # Crear perfil de cliente
            cliente = Cliente.objects.create(
                user=user,
                segundo_nombre=data.get('segundo_nombre', '').strip(),
                segundo_apellido=data.get('segundo_apellido', '').strip(),
                tipo_documento=data['tipo_documento'],
                numero_documento=numero_documento,
                telefono=telefono,
                compania=compania
            )
        
        return JsonResponse({
            'success': True,
            'message': 'El registro fue exitoso',
            'data': {
                'cliente_id': cliente.id,
                'nombre_completo': cliente.get_nombre_completo(),
                'email': email
            }
        }, status=201)
    
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error en el formato de los datos'
        }, status=400)
    
    except IntegrityError as e:
        return JsonResponse({
            'success': False,
            'message': 'Error de integridad en la base de datos'
        }, status=500)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}'
        }, status=500)


@admin_required
@csrf_exempt
@require_http_methods(["GET"])
def client_list_api(request):
    """
    Endpoint GET para consultar clientes.
    Permite filtrar por: email, numero_documento, compania_id
    
    Query params:
    - email: filtrar por email
    - numero_documento: filtrar por número de documento
    - compania_id: filtrar por compañía
    - activo: filtrar por estado (true/false)
    
    Requiere permisos de administrador (validación básica).
    """
    try:
        # TODO: Implementar validación de permisos de administrador
        # Por ahora permitimos la consulta
        
        # Obtener parámetros de filtro
        email = request.GET.get('email')
        numero_documento = request.GET.get('numero_documento')
        compania_id = request.GET.get('compania_id')
        activo = request.GET.get('activo')
        
        # Iniciar queryset
        clientes = Cliente.objects.select_related('user', 'compania').all()
        
        # Aplicar filtros
        if email:
            clientes = clientes.filter(user__email__icontains=email)
        
        if numero_documento:
            clientes = clientes.filter(numero_documento=numero_documento)
        
        if compania_id:
            clientes = clientes.filter(compania_id=compania_id)
        
        if activo is not None:
            activo_bool = activo.lower() == 'true'
            clientes = clientes.filter(activo=activo_bool)
        
        # Serializar datos (sin incluir contraseñas)
        clientes_data = []
        for cliente in clientes:
            clientes_data.append({
                'id': cliente.id,
                'nombre_completo': cliente.get_nombre_completo(),
                'primer_nombre': cliente.user.first_name,
                'segundo_nombre': cliente.segundo_nombre,
                'primer_apellido': cliente.user.last_name,
                'segundo_apellido': cliente.segundo_apellido,
                'tipo_documento': cliente.tipo_documento,
                'numero_documento': cliente.numero_documento,
                'email': cliente.user.email,
                'telefono': cliente.telefono,
                'compania': {
                    'id': cliente.compania.id,
                    'nombre': cliente.compania.nombre
                },
                'activo': cliente.activo,
                'fecha_registro': cliente.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return JsonResponse({
            'success': True,
            'count': len(clientes_data),
            'data': clientes_data
        }, status=200)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al consultar clientes: {str(e)}'
        }, status=500)


@admin_required
@csrf_exempt
@require_http_methods(["GET"])
def companies_list_api(request):
    """
    Endpoint GET para obtener lista de compañías activas.
    Usado para popular el select del formulario de registro.
    """
    try:
        companias = Compania.objects.filter(estado=True).order_by('nombre')
        
        companias_data = [
            {
                'id': c.id,
                'nombre': c.nombre
            }
            for c in companias
        ]
        
        return JsonResponse({
            'success': True,
            'data': companias_data
        }, status=200)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al consultar compañías: {str(e)}'
        }, status=500)


# ====================================
# MÓDULO: REGISTRO DE CONDUCTORES
# ====================================

@csrf_exempt
@require_http_methods(["POST"])
def driver_registration_api(request):
    """
    Endpoint POST para registrar nuevos conductores.
    Valida datos, verifica duplicados y crea el usuario + perfil de conductor + vehículo + documentos.
    
    Respuesta JSON:
    - success: true/false
    - message: mensaje descriptivo
    - data: información adicional (opcional)
    """
    try:
        # Parsear datos del formulario (multipart/form-data)
        data = request.POST
        files = request.FILES
        
        # ===================================
        # VALIDACIÓN DE CAMPOS OBLIGATORIOS
        # ===================================
        required_fields = [
            'primer_nombre', 'primer_apellido', 'tipo_documento',
            'numero_documento', 'fecha_nacimiento', 'correo', 
            'telefono_principal', 'direccion', 'ciudad', 'password',
            'numero_licencia', 'licencia_expedicion', 'licencia_vencimiento',
            'tipo_cuenta', 'banco', 'numero_cuenta', 'confirmar_numero_cuenta',
            'placa', 'marca', 'modelo', 'anio', 'color', 'tipo_vehiculo', 'num_pasajeros'
        ]
        
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return JsonResponse({
                'success': False,
                'message': 'Error: faltan campos obligatorios',
                'missing_fields': missing_fields
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE FORMATO DE NOMBRES
        # ===================================
        name_pattern = re.compile(r'^[A-Za-zÁÉÍÓÚáéíóúÑñ\s]+$')
        
        if not name_pattern.match(data['primer_nombre']):
            return JsonResponse({
                'success': False,
                'message': 'El primer nombre no debe contener números ni símbolos'
            }, status=400)
        
        if not name_pattern.match(data['primer_apellido']):
            return JsonResponse({
                'success': False,
                'message': 'El primer apellido no debe contener números ni símbolos'
            }, status=400)
        
        if data.get('segundo_nombre') and not name_pattern.match(data['segundo_nombre']):
            return JsonResponse({
                'success': False,
                'message': 'El segundo nombre no debe contener números ni símbolos'
            }, status=400)
        
        if data.get('segundo_apellido') and not name_pattern.match(data['segundo_apellido']):
            return JsonResponse({
                'success': False,
                'message': 'El segundo apellido no debe contener números ni símbolos'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE EMAIL
        # ===================================
        email = data['correo'].lower().strip()
        try:
            validate_email(email)
        except ValidationError:
            return JsonResponse({
                'success': False,
                'message': 'Verificar correo'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE CONTRASEÑA
        # ===================================
        password = data['password']
        confirm_password = data.get('confirm_password', '')
        
        if password != confirm_password:
            return JsonResponse({
                'success': False,
                'message': 'Las contraseñas no coinciden'
            }, status=400)
        
        # Validar requisitos de contraseña
        if len(password) < 8:
            return JsonResponse({
                'success': False,
                'message': 'La contraseña debe tener mínimo 8 caracteres'
            }, status=400)
        
        if not re.search(r'[A-Z]', password):
            return JsonResponse({
                'success': False,
                'message': 'La contraseña debe contener al menos una mayúscula'
            }, status=400)
        
        if not re.search(r'[a-z]', password):
            return JsonResponse({
                'success': False,
                'message': 'La contraseña debe contener al menos una minúscula'
            }, status=400)
        
        if not re.search(r'\d', password):
            return JsonResponse({
                'success': False,
                'message': 'La contraseña debe contener al menos un número'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE FECHA DE NACIMIENTO (EDAD MÍNIMA 21 AÑOS)
        # ===================================
        from datetime import date, datetime
        fecha_nacimiento = datetime.strptime(data['fecha_nacimiento'], '%Y-%m-%d').date()
        today = date.today()
        age = today.year - fecha_nacimiento.year - ((today.month, today.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
        
        if age < 21:
            return JsonResponse({
                'success': False,
                'message': 'El conductor debe ser mayor de 21 años'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE NÚMERO DE DOCUMENTO
        # ===================================
        numero_documento = data['numero_documento'].strip()
        if not numero_documento.isdigit():
            return JsonResponse({
                'success': False,
                'message': 'El número de documento solo debe contener números'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE TELÉFONOS
        # ===================================
        telefono_principal = data['telefono_principal'].strip()
        if not re.match(r'^[0-9]{10}$', telefono_principal):
            return JsonResponse({
                'success': False,
                'message': 'El teléfono principal debe tener 10 dígitos'
            }, status=400)
        
        telefono_secundario = data.get('telefono_secundario', '').strip()
        if telefono_secundario and not re.match(r'^[0-9]{10}$', telefono_secundario):
            return JsonResponse({
                'success': False,
                'message': 'El teléfono secundario debe tener 10 dígitos'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE NÚMEROS DE CUENTA
        # ===================================
        numero_cuenta = data['numero_cuenta'].strip()
        confirmar_numero_cuenta = data['confirmar_numero_cuenta'].strip()
        
        if numero_cuenta != confirmar_numero_cuenta:
            return JsonResponse({
                'success': False,
                'message': 'Los números de cuenta no coinciden'
            }, status=400)
        
        if not numero_cuenta.isdigit():
            return JsonResponse({
                'success': False,
                'message': 'El número de cuenta solo debe contener números'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE PLACA
        # ===================================
        placa = data['placa'].strip().upper()
        if not re.match(r'^[A-Z0-9]{1,6}$', placa):
            return JsonResponse({
                'success': False,
                'message': 'La placa solo puede tener letras y números (máximo 6 caracteres)'
            }, status=400)

        letras = sum(c.isalpha() for c in placa)
        numeros = sum(c.isdigit() for c in placa)

        if letras < 2 and numeros < 2:
            return JsonResponse({
                'success': False,
                'message': 'La placa debe tener al menos dos letras o dos números'
            }, status=400)
            
        # ===================================
        # VERIFICAR USUARIO NO EXISTENTE
        # ===================================
        if User.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'message': 'El usuario se encuentra registrado'
            }, status=409)
        
        if Conductor.objects.filter(numero_documento=numero_documento).exists():
            return JsonResponse({
                'success': False,
                'message': 'El conductor se encuentra registrado'
            }, status=409)
        
        if Vehiculo.objects.filter(placa=placa).exists():
            return JsonResponse({
                'success': False,
                'message': 'La placa del vehículo ya está registrada'
            }, status=409)
        
        # ===================================
        # VALIDACIÓN DE FECHAS DE LICENCIA
        # ===================================
        licencia_expedicion = datetime.strptime(data['licencia_expedicion'], '%Y-%m-%d').date()
        licencia_vencimiento = datetime.strptime(data['licencia_vencimiento'], '%Y-%m-%d').date()
        
        if licencia_vencimiento <= licencia_expedicion:
            return JsonResponse({
                'success': False,
                'message': 'La fecha de vencimiento debe ser posterior a la fecha de expedición'
            }, status=400)
        
        if licencia_vencimiento <= today:
            return JsonResponse({
                'success': False,
                'message': 'La licencia de conducción está vencida'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE AÑO DEL VEHÍCULO
        # ===================================
        anio_vehiculo = int(data['anio'])
        if anio_vehiculo < 2015 or anio_vehiculo > today.year:
            return JsonResponse({
                'success': False,
                'message': 'El año del vehículo debe estar entre 2015 y el año actual'
            }, status=400)
        
        # ===================================
        # VALIDACIÓN DE ARCHIVOS OBLIGATORIOS
        # ===================================
        required_files = [
            'documento_frontal', 'tarjeta_propiedad', 'certificado_reconocimiento',
            'foto_licencia', 'documento_soat', 'antecedentes_judiciales',
            'foto_vehiculo_frontal', 'foto_vehiculo_lateral', 'foto_vehiculo_interior',
            'certificado_tecnomecanica'
        ]
        
        missing_files = [file for file in required_files if file not in files]
        if missing_files:
            return JsonResponse({
                'success': False,
                'message': 'Faltan archivos obligatorios',
                'missing_files': missing_files
            }, status=400)
        
        # ===================================
        # CREAR USUARIO, CONDUCTOR, VEHÍCULO Y DOCUMENTOS
        # ===================================
        with transaction.atomic():
            # Crear usuario de Django
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=data['primer_nombre'].strip(),
                last_name=data['primer_apellido'].strip()
            )
            
            # Crear perfil de conductor
            conductor = Conductor.objects.create(
                user=user,
                segundo_nombre=data.get('segundo_nombre', '').strip() or None,
                segundo_apellido=data.get('segundo_apellido', '').strip() or None,
                tipo_documento=data['tipo_documento'],
                numero_documento=numero_documento,
                fecha_nacimiento=fecha_nacimiento,
                telefono_principal=telefono_principal,
                telefono_secundario=telefono_secundario or None,
                direccion=data['direccion'].strip(),
                ciudad=data['ciudad'],
                numero_licencia=data['numero_licencia'].strip(),
                licencia_expedicion=licencia_expedicion,
                licencia_vencimiento=licencia_vencimiento,
                tipo_cuenta=data['tipo_cuenta'],
                banco=data['banco'],
                numero_cuenta=numero_cuenta
            )
            
            # Crear vehículo
            vehiculo = Vehiculo.objects.create(
                conductor=conductor,
                placa=placa,
                marca=data['marca'].strip().upper(),
                modelo=data['modelo'].strip(),
                anio=anio_vehiculo,
                color=data['color'].strip(),
                tipo_vehiculo=data['tipo_vehiculo'].strip(),
                num_pasajeros=int(data['num_pasajeros'])
            )
            
            # Crear documentos
            documento_types = {
                'documento_frontal': 'documento_frontal',
                'documento_reverso': 'documento_reverso',
                'tarjeta_propiedad': 'tarjeta_propiedad',
                'certificado_reconocimiento': 'certificado_reconocimiento',
                'foto_licencia': 'foto_licencia',
                'documento_soat': 'documento_soat',
                'antecedentes_judiciales': 'antecedentes_judiciales',
                'foto_vehiculo_frontal': 'foto_vehiculo_frontal',
                'foto_vehiculo_lateral': 'foto_vehiculo_lateral',
                'foto_vehiculo_interior': 'foto_vehiculo_interior',
                'certificado_tecnomecanica': 'certificado_tecnomecanica'
            }
            
            for file_field, tipo_doc in documento_types.items():
                if file_field in files:
                    file = files[file_field]
                    DocumentoConductor.objects.create(
                        conductor=conductor,
                        tipo_documento=tipo_doc,
                        archivo=file,
                        nombre_original=file.name,
                        tamaño_archivo=file.size
                    )
        
        return JsonResponse({
            'success': True,
            'message': 'El registro del conductor fue exitoso. Estado: Pendiente de Verificación',
            'data': {
                'conductor_id': conductor.id,
                'nombre_completo': conductor.get_nombre_completo(),
                'email': email,
                'estado': conductor.estado,
                'vehiculo_placa': vehiculo.placa
            }
        }, status=201)
    
    except IntegrityError as e:
        return JsonResponse({
            'success': False,
            'message': 'Error de integridad en la base de datos'
        }, status=500)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def driver_list_api(request):
    """
    Endpoint GET para consultar conductores.
    Permite filtrar por: email, numero_documento, estado, placa
    """
    try:
        # Obtener parámetros de filtro
        email = request.GET.get('email')
        numero_documento = request.GET.get('numero_documento')
        estado = request.GET.get('estado')
        placa = request.GET.get('placa')
        activo = request.GET.get('activo')
        
        # Iniciar queryset
        conductores = Conductor.objects.select_related('user', 'vehiculo').all()
        
        # Aplicar filtros
        if email:
            conductores = conductores.filter(user__email__icontains=email)
        
        if numero_documento:
            conductores = conductores.filter(numero_documento=numero_documento)
        
        if estado:
            conductores = conductores.filter(estado=estado)
        
        if placa:
            conductores = conductores.filter(vehiculo__placa__icontains=placa)
        
        if activo is not None:
            activo_bool = activo.lower() == 'true'
            conductores = conductores.filter(activo=activo_bool)
        
        # Serializar datos
        conductores_data = []
        for conductor in conductores:
            vehiculo_data = None
            if hasattr(conductor, 'vehiculo'):
                vehiculo_data = {
                    'placa': conductor.vehiculo.placa,
                    'marca': conductor.vehiculo.marca,
                    'modelo': conductor.vehiculo.modelo,
                    'anio': conductor.vehiculo.anio,
                    'color': conductor.vehiculo.color,
                    'tipo_vehiculo': conductor.vehiculo.tipo_vehiculo,
                    'num_pasajeros': conductor.vehiculo.num_pasajeros
                }
            
            conductores_data.append({
                'id': conductor.id,
                'nombre_completo': conductor.get_nombre_completo(),
                'primer_nombre': conductor.user.first_name,
                'segundo_nombre': conductor.segundo_nombre,
                'primer_apellido': conductor.user.last_name,
                'segundo_apellido': conductor.segundo_apellido,
                'tipo_documento': conductor.tipo_documento,
                'numero_documento': conductor.numero_documento,
                'fecha_nacimiento': conductor.fecha_nacimiento.strftime('%Y-%m-%d'),
                'edad': conductor.get_edad(),
                'email': conductor.user.email,
                'telefono_principal': conductor.telefono_principal,
                'telefono_secundario': conductor.telefono_secundario,
                'direccion': conductor.direccion,
                'ciudad': conductor.ciudad,
                'numero_licencia': conductor.numero_licencia,
                'licencia_expedicion': conductor.licencia_expedicion.strftime('%Y-%m-%d'),
                'licencia_vencimiento': conductor.licencia_vencimiento.strftime('%Y-%m-%d'),
                'tipo_cuenta': conductor.tipo_cuenta,
                'banco': conductor.banco,
                'numero_cuenta': conductor.numero_cuenta,
                'estado': conductor.estado,
                'vehiculo': vehiculo_data,
                'activo': conductor.activo,
                'fecha_registro': conductor.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return JsonResponse({
            'success': True,
            'count': len(conductores_data),
            'data': conductores_data
        }, status=200)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al consultar conductores: {str(e)}'
        }, status=500)


# ====================================
# MÓDULO: DETALLE DE COMPAÑÍAS
# ====================================

@login_required
def companies_list(request):
    """
    Vista que muestra el listado de todas las compañías.
    GET: Muestra la lista de compañías disponibles
    """
    companias = Compania.objects.filter(estado=True).order_by('nombre')
    return render(request, 'companies_list.html', {'companias': companias})


@login_required
def company_detail(request, company_id):
    """
    Vista que renderiza el módulo de detalle de compañía.
    GET: Muestra el dashboard con pestañas
    """
    try:
        compania = Compania.objects.get(id=company_id)
        return render(request, 'company_detail.html', {'compania': compania})
    except Compania.DoesNotExist:
        messages.error(request, 'La compañía no existe')
        return redirect('inicio')


@csrf_exempt
@require_http_methods(["GET"])
def company_search_api(request):
    """
    Endpoint GET para buscar compañías.
    Query params:
    - q: término de búsqueda (nombre, NIT, razón social)
    - estado: filtrar por estado de cuenta
    """
    try:
        query = request.GET.get('q', '').strip()
        estado = request.GET.get('estado')
        
        companias = Compania.objects.filter(estado=True)
        
        if query:
            from django.db.models import Q
            companias = companias.filter(
                Q(nombre__icontains=query) |
                Q(nit__icontains=query) |
                Q(razon_social__icontains=query)
            )
        
        if estado:
            companias = companias.filter(estado_cuenta=estado)
        
        companias_data = []
        for c in companias:
            companias_data.append({
                'id': c.id,
                'nombre': c.nombre,
                'razon_social': c.razon_social,
                'nit': c.nit,
                'estado_cuenta': c.estado_cuenta,
            })
        
        return JsonResponse({
            'success': True,
            'count': len(companias_data),
            'data': companias_data
        }, status=200)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error en la búsqueda: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def company_detail_api(request, company_id):
    """
    Endpoint GET para obtener detalle completo de una compañía con métricas.
    Incluye: información básica, métricas de servicios, empleados activos, etc.
    """
    try:
        compania = Compania.objects.get(id=company_id)
        
        # Calcular métricas
        from datetime import datetime, timedelta
        from django.db.models import Count, Sum, Q
        
        # Total de clientes activos
        clientes_activos = Cliente.objects.filter(compania=compania, activo=True).count()
        
        # Servicios del mes actual
        hoy = datetime.now()
        primer_dia_mes = hoy.replace(day=1)
        servicios_mes = Viaje.objects.filter(
            cliente__compania=compania,
            fecha_solicitud__gte=primer_dia_mes
        ).count()
        
        # Total de servicios realizados (TODOS los viajes, no solo completados)
        servicios_realizados = Viaje.objects.filter(
            cliente__compania=compania
        ).count()
        
        # Calcular porcentaje del mes actual vs mes anterior
        mes_anterior_inicio = (primer_dia_mes - timedelta(days=1)).replace(day=1)
        servicios_mes_anterior = Viaje.objects.filter(
            cliente__compania=compania,
            fecha_solicitud__gte=mes_anterior_inicio,
            fecha_solicitud__lt=primer_dia_mes
        ).count()
        
        porcentaje_mes = 0
        if servicios_mes_anterior > 0:
            porcentaje_mes = ((servicios_mes - servicios_mes_anterior) / servicios_mes_anterior) * 100
        elif servicios_mes > 0:
            porcentaje_mes = 100
        
        compania_data = {
            'id': compania.id,
            'nombre': compania.nombre,
            'razon_social': compania.razon_social,
            'nit': compania.nit,
            'direccion': compania.direccion,
            'telefono': compania.telefono,
            'email_corporativo': compania.email_corporativo,
            'persona_contacto': compania.persona_contacto,
            'fecha_membresia': compania.fecha_membresia.strftime('%Y-%m-%d') if compania.fecha_membresia else '',
            'estado_cuenta': compania.estado_cuenta,
            'estado': compania.estado,
            # Métricas
            'metricas': {
                'servicios_realizados': servicios_realizados,
                'empleados_activos': clientes_activos,
                'servicios_mes': servicios_mes,
                'porcentaje_mes': round(porcentaje_mes, 1)
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': compania_data
        }, status=200)
    
    except Compania.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Compañía no encontrada'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al consultar la compañía: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def company_clients_api(request, company_id):
    """
    Endpoint GET para obtener clientes asociados a una compañía.
    Incluye: nombre, ID, cargo, total viajes, calificación promedio, último viaje, estado.
    """
    try:
        compania = Compania.objects.get(id=company_id)
        clientes = Cliente.objects.filter(compania=compania).select_related('user')
        
        from django.db.models import Count, Avg, Max
        
        clientes_data = []
        for cliente in clientes:
            # Calcular total de viajes
            total_viajes = Viaje.objects.filter(cliente=cliente).count()
            
            # Calcular calificación promedio del cliente
            calificacion_promedio = Viaje.objects.filter(
                cliente=cliente,
                calificacion_cliente__isnull=False
            ).aggregate(Avg('calificacion_cliente'))['calificacion_cliente__avg']
            
            # Obtener último viaje
            ultimo_viaje = Viaje.objects.filter(cliente=cliente).order_by('-fecha_solicitud').first()
            ultimo_viaje_data = None
            if ultimo_viaje:
                ultimo_viaje_data = {
                    'fecha': ultimo_viaje.fecha_solicitud.strftime('%Y-%m-%d %H:%M'),
                    'origen': ultimo_viaje.origen,
                    'destino': ultimo_viaje.destino,
                    'estado': ultimo_viaje.estado
                }
            
            clientes_data.append({
                'id': cliente.id,
                'nombre_completo': cliente.get_nombre_completo(),
                'cargo': cliente.cargo or 'Sin cargo',
                'numero_documento': cliente.numero_documento,
                'email': cliente.user.email,
                'telefono': cliente.telefono,
                'total_viajes': total_viajes,
                'calificacion_promedio': round(calificacion_promedio, 1) if calificacion_promedio else 0,
                'ultimo_viaje': ultimo_viaje_data,
                'activo': cliente.activo,
                'fecha_registro': cliente.fecha_registro.strftime('%Y-%m-%d')
            })
        
        return JsonResponse({
            'success': True,
            'count': len(clientes_data),
            'data': clientes_data
        }, status=200)
    
    except Compania.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Compañía no encontrada'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al consultar clientes: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["PATCH"])
def client_toggle_status_api(request, client_id):
    """
    Endpoint PATCH para activar/desactivar un cliente.
    Body JSON:
    - activo: true/false
    """
    try:
        data = json.loads(request.body)
        cliente = Cliente.objects.get(id=client_id)
        
        nuevo_estado = data.get('activo')
        if nuevo_estado is None:
            return JsonResponse({
                'success': False,
                'message': 'El campo "activo" es requerido'
            }, status=400)
        
        cliente.activo = nuevo_estado
        cliente.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Cliente {"activado" if nuevo_estado else "desactivado"} correctamente',
            'data': {
                'id': cliente.id,
                'nombre_completo': cliente.get_nombre_completo(),
                'activo': cliente.activo
            }
        }, status=200)
    
    except Cliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Cliente no encontrado'
        }, status=404)
    
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error en el formato de los datos'
        }, status=400)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al actualizar el cliente: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def generate_services_report_api(request):
    """
    Endpoint POST para generar reporte de servicios solicitados.
    Body JSON:
    - company_id: ID de la compañía
    - fecha_inicio: fecha de inicio (YYYY-MM-DD)
    - fecha_fin: fecha de fin (YYYY-MM-DD)
    - export: true para exportar a Excel, false para solo JSON
    """
    try:
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        data = json.loads(request.body)
        company_id = data.get('company_id')
        fecha_inicio_str = data.get('fecha_inicio')
        fecha_fin_str = data.get('fecha_fin')
        export = data.get('export', False)
        
        if not all([company_id, fecha_inicio_str, fecha_fin_str]):
            return JsonResponse({
                'success': False,
                'message': 'Faltan campos obligatorios'
            }, status=400)
        
        compania = Compania.objects.get(id=company_id)
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
        
        # Consultar viajes en el rango de fechas
        viajes = Viaje.objects.filter(
            cliente__compania=compania,
            fecha_solicitud__gte=fecha_inicio,
            fecha_solicitud__lte=fecha_fin
        ).select_related('cliente__user', 'conductor__user', 'cliente__compania')
        
        reportes_data = []
        for viaje in viajes:
            reportes_data.append({
                'id_cliente': viaje.cliente.id,
                'nombre_cliente': viaje.cliente.get_nombre_completo(),
                'nombre_empresa': viaje.cliente.compania.nombre,
                'id_empresa': viaje.cliente.compania.id,
                'fecha': viaje.fecha_solicitud.strftime('%Y-%m-%d %H:%M'),
                'id_conductor': viaje.conductor.id,
                'nombre_conductor': viaje.conductor.get_nombre_completo(),
                'origen': viaje.origen,
                'destino': viaje.destino,
                'estado': viaje.estado,
                'valor_total': str(viaje.valor_total)
            })
        
        # Si se solicita exportación a Excel
        if export:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Servicios"
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Encabezados
            headers = ['ID Cliente', 'Nombre Cliente', 'Empresa', 'ID Empresa', 'Fecha', 
                      'ID Conductor', 'Nombre Conductor', 'Origen', 'Destino', 'Estado', 'Valor Total']
            ws.append(headers)
            
            # Aplicar estilo a encabezados
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos
            for reporte in reportes_data:
                ws.append([
                    reporte['id_cliente'],
                    reporte['nombre_cliente'],
                    reporte['nombre_empresa'],
                    reporte['id_empresa'],
                    reporte['fecha'],
                    reporte['id_conductor'],
                    reporte['nombre_conductor'],
                    reporte['origen'],
                    reporte['destino'],
                    reporte['estado'],
                    reporte['valor_total']
                ])
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Preparar respuesta de descarga
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=reporte_servicios_{compania.nombre}_{fecha_inicio_str}_{fecha_fin_str}.xlsx'
            wb.save(response)
            return response
        
        # Respuesta JSON
        return JsonResponse({
            'success': True,
            'count': len(reportes_data),
            'data': reportes_data
        }, status=200)
    
    except Compania.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Compañía no encontrada'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al generar reporte: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def generate_income_report_api(request):
    """
    Endpoint POST para generar reporte de ingresos.
    Body JSON:
    - company_id: ID de la compañía
    - fecha_inicio: fecha de inicio (YYYY-MM-DD)
    - fecha_fin: fecha de fin (YYYY-MM-DD)
    - export: true para exportar a Excel
    """
    try:
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, numbers
        from django.http import HttpResponse
        from django.db.models import Sum
        
        data = json.loads(request.body)
        company_id = data.get('company_id')
        fecha_inicio_str = data.get('fecha_inicio')
        fecha_fin_str = data.get('fecha_fin')
        export = data.get('export', False)
        
        if not all([company_id, fecha_inicio_str, fecha_fin_str]):
            return JsonResponse({
                'success': False,
                'message': 'Faltan campos obligatorios'
            }, status=400)
        
        compania = Compania.objects.get(id=company_id)
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
        
        # Consultar viajes completados en el rango de fechas
        viajes = Viaje.objects.filter(
            cliente__compania=compania,
            fecha_solicitud__gte=fecha_inicio,
            fecha_solicitud__lte=fecha_fin,
            estado='Completado'
        ).select_related('cliente__compania')
        
        reportes_data = []
        total_ingresos = 0
        
        for viaje in viajes:
            monto = float(viaje.valor_total)
            total_ingresos += monto
            
            reportes_data.append({
                'id_viaje': viaje.id,
                'fecha': viaje.fecha_solicitud.strftime('%Y-%m-%d %H:%M'),
                'id_empresa': viaje.cliente.compania.id,
                'nombre_empresa': viaje.cliente.compania.nombre,
                'monto': monto,
                'metodo_pago': viaje.metodo_pago,
                'origen': viaje.origen,
                'destino': viaje.destino
            })
        
        # Si se solicita exportación a Excel
        if export:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Ingresos"
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Título
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = f"REPORTE DE INGRESOS - {compania.nombre}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Encabezados
            ws.append([])
            headers = ['ID Viaje', 'Fecha', 'ID Empresa', 'Nombre Empresa', 'Monto', 'Método de Pago', 'Origen', 'Destino']
            ws.append(headers)
            
            # Aplicar estilo a encabezados
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos
            for reporte in reportes_data:
                ws.append([
                    reporte['id_viaje'],
                    reporte['fecha'],
                    reporte['id_empresa'],
                    reporte['nombre_empresa'],
                    reporte['monto'],
                    reporte['metodo_pago'],
                    reporte['origen'],
                    reporte['destino']
                ])
            
            # Fila de total
            ws.append([])
            total_row = ws.max_row + 1
            ws[f'D{total_row}'] = 'TOTAL INGRESOS:'
            ws[f'D{total_row}'].font = Font(bold=True)
            ws[f'E{total_row}'] = total_ingresos
            ws[f'E{total_row}'].font = Font(bold=True, color="00AA00")
            ws[f'E{total_row}'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 25
            
            # Preparar respuesta de descarga
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=reporte_ingresos_{compania.nombre}_{fecha_inicio_str}_{fecha_fin_str}.xlsx'
            wb.save(response)
            return response
        
        # Respuesta JSON
        return JsonResponse({
            'success': True,
            'count': len(reportes_data),
            'total_ingresos': total_ingresos,
            'data': reportes_data
        }, status=200)
    
    except Compania.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Compañía no encontrada'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al generar reporte: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def generate_issues_report_api(request):
    """
    Endpoint POST para generar reporte de novedades.
    Body JSON:
    - company_id: ID de la compañía
    - fecha_inicio: fecha de inicio (YYYY-MM-DD)
    - fecha_fin: fecha de fin (YYYY-MM-DD)
    - export: true para exportar a Excel
    """
    try:
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        data = json.loads(request.body)
        company_id = data.get('company_id')
        fecha_inicio_str = data.get('fecha_inicio')
        fecha_fin_str = data.get('fecha_fin')
        export = data.get('export', False)
        
        if not all([company_id, fecha_inicio_str, fecha_fin_str]):
            return JsonResponse({
                'success': False,
                'message': 'Faltan campos obligatorios'
            }, status=400)
        
        compania = Compania.objects.get(id=company_id)
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
        
        # Consultar novedades en el rango de fechas
        novedades = Novedad.objects.filter(
            compania=compania,
            fecha_creacion__gte=fecha_inicio,
            fecha_creacion__lte=fecha_fin
        ).select_related('creado_por', 'viaje', 'conductor__user', 'cliente__user')
        
        reportes_data = []
        for novedad in novedades:
            reportes_data.append({
                'id_novedad': novedad.id,
                'fecha_creacion': novedad.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
                'tipo_novedad': novedad.tipo_novedad,
                'descripcion': novedad.descripcion,
                'estado': novedad.estado,
                'prioridad': novedad.prioridad,
                'creado_por': novedad.creado_por.get_full_name() if novedad.creado_por else 'Sistema',
                'id_viaje': novedad.viaje.id if novedad.viaje else 'N/A',
                'conductor': novedad.conductor.get_nombre_completo() if novedad.conductor else 'N/A',
                'cliente': novedad.cliente.get_nombre_completo() if novedad.cliente else 'N/A',
                'fecha_resolucion': novedad.fecha_resolucion.strftime('%Y-%m-%d %H:%M') if novedad.fecha_resolucion else 'Pendiente'
            })
        
        # Si se solicita exportación a Excel
        if export:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Novedades"
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Título
            ws.merge_cells('A1:K1')
            title_cell = ws['A1']
            title_cell.value = f"REPORTE DE NOVEDADES - {compania.nombre}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Encabezados
            ws.append([])
            headers = ['ID', 'Fecha Creación', 'Tipo', 'Descripción', 'Estado', 'Prioridad', 
                      'Creado Por', 'ID Viaje', 'Conductor', 'Cliente', 'Fecha Resolución']
            ws.append(headers)
            
            # Aplicar estilo a encabezados
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos
            for reporte in reportes_data:
                ws.append([
                    reporte['id_novedad'],
                    reporte['fecha_creacion'],
                    reporte['tipo_novedad'],
                    reporte['descripcion'],
                    reporte['estado'],
                    reporte['prioridad'],
                    reporte['creado_por'],
                    reporte['id_viaje'],
                    reporte['conductor'],
                    reporte['cliente'],
                    reporte['fecha_resolucion']
                ])
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 25
            ws.column_dimensions['J'].width = 25
            ws.column_dimensions['K'].width = 18
            
            # Preparar respuesta de descarga
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=reporte_novedades_{compania.nombre}_{fecha_inicio_str}_{fecha_fin_str}.xlsx'
            wb.save(response)
            return response
        
        # Respuesta JSON
        return JsonResponse({
            'success': True,
            'count': len(reportes_data),
            'data': reportes_data
        }, status=200)
    
    except Compania.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Compañía no encontrada'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al generar reporte: {str(e)}'
        }, status=500)


def custom_404_view(request, exception):
    from django.shortcuts import redirect
    return redirect('home')

