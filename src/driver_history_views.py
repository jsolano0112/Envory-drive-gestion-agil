# views/driver_history_views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Avg, Count
from datetime import datetime, date
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Conductor, Vehiculo


@login_required
def driver_history(request):
    """Vista principal del historial de conductor"""
    context = {
        'driver': None,
        'search_term': '',
        'active_tab': 'viajes',
        'date_start': '',
        'date_end': ''
    }
    
    search_term = request.GET.get('search', '')
    active_tab = request.GET.get('tab', 'viajes')
    date_start = request.GET.get('date_start', '')
    date_end = request.GET.get('date_end', '')
    
    if search_term:
        # Buscar conductor por nombre, documento o placa
        driver = Conductor.objects.filter(
            Q(user__first_name__icontains=search_term) |
            Q(user__last_name__icontains=search_term) |
            Q(numero_documento__icontains=search_term) |
            Q(vehiculo__placa__icontains=search_term)
        ).select_related('user', 'vehiculo').first()
        
        if driver:
            stats = {
                'total_trips': 247,
                'avg_rating': 4.8,
                'completion_rate': 96,
                'total_reports': 2
            }
                        # Data mockeada de viajes realista
            trips = []
            base_date = datetime.now()

            for i in range(5):
                # Fechas aleatorias en los últimos 30 días
                days_ago = random.randint(0, 30)
                trip_date = base_date - timedelta(days=days_ago, hours=random.randint(6, 22))
                
                # Estados posibles con probabilidades realistas
                status_weights = [('Completado', 0.85), ('Cancelado', 0.10), ('En Progreso', 0.04), ('Aceptado', 0.01)]
                status = random.choices([s[0] for s in status_weights], weights=[s[1] for s in status_weights])[0]
                
                # Calificación solo para viajes completados
                rating = random.choice([4.5, 5.0, 4.0, 4.5, 3.5, 5.0, 4.0, 4.5, 5.0, 3.0]) if status == 'Completado' and random.random() > 0.1 else None
                
                # Ubicaciones realistas en Bogotá
                locations = [
                    'Centro Internacional', 'Aeropuerto El Dorado', 'Zona Rosa', 'Parque de la 93',
                    'Centro Andino', 'Gran Estación', 'Plaza de las Américas', 'Calle 80',
                    'Centro Empresarial Santa Fe', 'Parque Nacional', 'Usaquén', 'Chapinero',
                    'Salitre Plaza', 'Portal Norte', 'Calle 170', 'Centro Histórico'
                ]
                
                origin = random.choice(locations)
                destination = random.choice([loc for loc in locations if loc != origin])
                
                # Clientes mockeados
                clients = [
                    'María González', 'Carlos Rodríguez', 'Ana López', 'Juan Martínez', 
                    'Laura Sánchez', 'Diego Pérez', 'Camila Torres', 'Andrés Ramírez',
                    'Valentina Morales', 'Santiago Herrera', 'Isabella Castro', 'Mateo Vargas'
                ]
                
                trip = {
                    'id': f'TRP{1000 + i}',
                    'date': trip_date,
                    'status': status,
                    'rating': rating,
                    'origin': origin,
                    'destination': destination,
                    'client': random.choice(clients),
                    'amount': random.randint(15000, 45000),
                    'distance': round(random.uniform(5.0, 25.0), 1),
                    'duration': f"{random.randint(20, 90):02d}:{random.randint(0, 59):02d}:00"
                }
                trips.append(trip)

            # Data mockeada de historial de estados realista
            status_history = []
            status_sequence = ['Pendiente', 'Aprobado', 'Activo', 'Suspendido', 'Aprobado', 'Activo']

            current_date = driver.fecha_registro
            admins = ['Admin Sistema', 'María González', 'Carlos Rodríguez', 'Supervisor General']

            for i, new_status in enumerate(status_sequence):
                if i == 0:  # Primer estado
                    prev_status = 'Pendiente'
                    title = 'Registro Inicial'
                    reason = 'Registro inicial del conductor en la plataforma'
                else:
                    prev_status = status_sequence[i-1]
                    transitions = {
                        ('Pendiente', 'Aprobado'): ('Conductor Aprobado', 'Documentos verificados correctamente'),
                        ('Aprobado', 'Activo'): ('Conductor Activado', 'Primer inicio de sesión exitoso'),
                        ('Activo', 'Suspendido'): ('Conductor Suspendido', 'Queja de cliente por demora excesiva'),
                        ('Suspendido', 'Aprobado'): ('Conductor Reactivado', 'Cumplimiento de condiciones de reactivación'),
                    }
                    title, reason = transitions.get((prev_status, new_status), ('Cambio de Estado', 'Cambio administrativo'))
                
                # Fechas progresivas
                days_later = random.randint(1, 30) if i > 0 else 0
                current_date += timedelta(days=days_later)
                
                change = {
                    'id': i + 1,
                    'date': current_date,
                    'title': title,
                    'previous_status': prev_status,
                    'new_status': new_status,
                    'reason': reason,
                    'modified_by': random.choice(admins)
                }
                status_history.append(change)

            # Estadísticas calculadas de la data mockeada
            total_trips = len(trips)
            completed_trips = len([t for t in trips if t['status'] == 'Completado'])
            avg_rating = sum(t['rating'] for t in trips if t['rating']) / len([t for t in trips if t['rating']]) if any(t['rating'] for t in trips) else 0
            completion_rate = int((completed_trips / total_trips * 100)) if total_trips > 0 else 0

            stats = {
                'total_trips': total_trips,
                'avg_rating': round(avg_rating, 1) if avg_rating > 0 else 4.8,
                'completion_rate': completion_rate,
                'total_reports': random.randint(0, 5)
            }
            
            context.update({
                'driver': driver,
                'stats': stats,
                'trips': trips,
                'status_history': status_history,
                'search_term': search_term,
                'active_tab': active_tab,
                'date_start': date_start,
                'date_end': date_end
            })
    
    return render(request, 'driver_details.html', context)


@csrf_exempt
@require_http_methods(["POST"])
def update_driver_status(request, driver_id):
    try:
        driver = get_object_or_404(Conductor, id=driver_id)
        new_status = request.POST.get('status')
        
        # Validar estado
        valid_statuses = ['Activo', 'Inactivo', 'Suspendido', 'Bloqueado', 'En Revisión']
        if new_status not in valid_statuses:
            return JsonResponse({
                'success': False,
                'message': 'Estado no válido'
            }, status=400)
        
        # Guardar estado anterior
        old_status = driver.estado
        
        # Actualizar estado del conductor
        driver.estado = new_status
        driver.save()
        
        # TODO: Crear registro en historial de cambios de estado
        # StatusHistory.objects.create(
        #     conductor=driver,
        #     previous_status=old_status,
        #     new_status=new_status,
        #     reason=request.POST.get('reason', 'Cambio manual'),
        #     modified_by=request.user
        # )
        
        # Redirigir de vuelta al historial con el mismo conductor
        return redirect(f"{request.META.get('HTTP_REFERER', '/')}?search={driver.numero_documento}")
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al actualizar estado: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def generate_report(request, driver_id):
    try:
        driver = get_object_or_404(Conductor, id=driver_id)
        report_type = request.POST.get('report_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Validar campos requeridos
        if not report_type or not start_date or not end_date:
            return JsonResponse({
                'success': False,
                'message': 'Faltan campos requeridos'
            }, status=400)
        
        # Convertir fechas
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Validar que la fecha de inicio sea anterior a la fecha final
        if start_date_obj > end_date_obj:
            return JsonResponse({
                'success': False,
                'message': 'La fecha de inicio debe ser anterior a la fecha final'
            }, status=400)
        
        # Crear workbook de Excel
        wb = Workbook()
        ws = wb.active
        
        # Estilos para el encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Información del reporte (primera fila)
        ws['A1'] = f"Reporte {report_type.title()}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:K1')
        
        # Información del conductor
        ws['A3'] = "Información del Conductor"
        ws['A3'].font = Font(bold=True)
        ws.merge_cells('A3:K3')
        
        conductor_info = [
            ["Nombre:", driver.get_nombre_completo()],
            ["Documento:", driver.numero_documento],
            ["Placa:", driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else 'N/A'],
            ["Período:", f"{start_date} al {end_date}"],
        ]
        
        for i, (label, value) in enumerate(conductor_info, 5):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = value
            ws.merge_cells(f'B{i}:K{i}')
        
        # Espacio en blanco
        current_row = len(conductor_info) + 6
        
        if report_type in ['mensual', 'diario']:
            # Reporte de viajes
            ws[f'A{current_row}'] = "REPORTE DE VIAJES"
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            ws.merge_cells(f'A{current_row}:K{current_row}')
            
            # Encabezados
            headers = [
                "ID Conductor", "Nombre Conductor", "Placa Vehículo", "ID Cliente", 
                "Nombre Cliente", "Lugar Recogida", "Destino", "Distancia (km)", 
                "Tiempo Acumulado", "Fecha Inicio", "Fecha Fin"
            ]
            
            header_row = current_row + 2
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = border
            
            # TODO: Cuando tengas el modelo Trip, reemplaza estos datos de ejemplo
            # trips = Trip.objects.filter(
            #     conductor=driver,
            #     fecha_inicio__date__range=[start_date_obj, end_date_obj]
            # ).select_related('cliente')
            
            # Datos de ejemplo
            sample_trips = [
                [driver.id, driver.get_nombre_completo(), driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else 'N/A', 
                 "CLI001", "Juan Pérez", "Centro", "Aeropuerto", 15.5, "02:30:00", "2025-01-15 08:00", "2025-01-15 10:30"],
                [driver.id, driver.get_nombre_completo(), driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else 'N/A',
                 "CLI002", "María González", "Mall", "Hospital", 8.2, "01:45:00", "2025-01-16 14:00", "2025-01-16 15:45"],
                [driver.id, driver.get_nombre_completo(), driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else 'N/A',
                 "CLI003", "Carlos Rodríguez", "Estadio", "Centro", 12.8, "02:15:00", "2025-01-17 19:00", "2025-01-17 21:15"],
            ]
            
            # Agregar datos
            for row_idx, trip in enumerate(sample_trips, header_row + 1):
                for col_idx, value in enumerate(trip, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
            
            # Total de viajes
            total_row = header_row + len(sample_trips) + 2
            ws[f'A{total_row}'] = f"TOTAL VIAJES: {len(sample_trips)}"
            ws[f'A{total_row}'].font = Font(bold=True)
            ws.merge_cells(f'A{total_row}:K{total_row}')
            
        elif report_type == 'monetario':
            # Reporte monetario
            ws[f'A{current_row}'] = "REPORTE MONETARIO"
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            ws.merge_cells(f'A{current_row}:K{current_row}')
            
            # Encabezados
            headers = ["ID Conductor", "Nombre Conductor", "Fecha Inicio", "Fecha Fin", "Valor Total Generado"]
            
            header_row = current_row + 2
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = border
            
            # TODO: Cuando tengas el modelo Trip, calcula el total real
            # total_value = Trip.objects.filter(
            #     conductor=driver,
            #     fecha_inicio__date__range=[start_date_obj, end_date_obj]
            # ).aggregate(total=Sum('valor'))['total'] or 0
            
            # Datos de ejemplo
            total_value = 2850000  # $2,850,000
            
            data_row = header_row + 1
            ws.cell(row=data_row, column=1).value = driver.id
            ws.cell(row=data_row, column=2).value = driver.get_nombre_completo()
            ws.cell(row=data_row, column=3).value = start_date
            ws.cell(row=data_row, column=4).value = end_date
            ws.cell(row=data_row, column=5).value = f"${total_value:,.0f}"
            
            # Aplicar bordes
            for col in range(1, 6):
                ws.cell(row=data_row, column=col).border = border
        
        # Ajustar ancho de columnas
        for col in range(1, 12):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Crear respuesta HTTP
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Preparar la respuesta
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'reporte_{report_type}_{driver.numero_documento}_{start_date}_{end_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al generar reporte: {str(e)}'
        }, status=500)


def export_history(request, driver_id):
    """
    Exporta el historial completo del conductor a CSV.
    Incluye: información personal, vehículo, estadísticas y viajes.
    """
    try:
        driver = get_object_or_404(Conductor, id=driver_id)
        
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="historial_{driver.numero_documento}.csv"'
        
        # Agregar BOM para Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Escribir información del conductor
        writer.writerow(['INFORMACIÓN DEL CONDUCTOR'])
        writer.writerow(['Nombre Completo', driver.get_nombre_completo()])
        writer.writerow(['Documento', driver.numero_documento])
        writer.writerow(['Email', driver.user.email])
        writer.writerow(['Teléfono', driver.telefono_principal])
        if driver.telefono_secundario:
            writer.writerow(['Teléfono Secundario', driver.telefono_secundario])
        writer.writerow(['Ciudad', driver.ciudad])
        writer.writerow(['Estado', driver.estado])
        writer.writerow(['Fecha Registro', driver.fecha_registro.strftime('%Y-%m-%d')])
        writer.writerow([])  # Línea vacía
        
        # Información del vehículo
        if hasattr(driver, 'vehiculo') and driver.vehiculo:
            writer.writerow(['INFORMACIÓN DEL VEHÍCULO'])
            writer.writerow(['Placa', driver.vehiculo.placa])
            writer.writerow(['Marca', driver.vehiculo.marca])
            writer.writerow(['Modelo', driver.vehiculo.modelo])
            writer.writerow(['Año', driver.vehiculo.anio])
            writer.writerow(['Color', driver.vehiculo.color])
            writer.writerow(['Tipo', driver.vehiculo.tipo_vehiculo])
            writer.writerow([])
        
        # Estadísticas
        writer.writerow(['ESTADÍSTICAS'])
        writer.writerow(['Total Viajes', '247'])  # TODO: Reemplazar con datos reales
        writer.writerow(['Calificación Promedio', '4.8'])
        writer.writerow(['Tasa Completado', '96%'])
        writer.writerow(['Reportes', '2'])
        writer.writerow([])
        
       
        return response
    
    except Exception as e:
        return HttpResponse(f'Error al exportar: {str(e)}', status=500)


@require_http_methods(["GET"])
def driver_statistics_api(request, driver_id):
    """
    Endpoint API para obtener estadísticas del conductor en formato JSON.
    Útil para dashboards o consultas programáticas.
    """
    try:
        driver = get_object_or_404(Conductor, id=driver_id)
        
        # TODO: Calcular con datos reales cuando tengas el modelo Trip
        # trips = Trip.objects.filter(conductor=driver)
        # total_trips = trips.count()
        # completed_trips = trips.filter(estado='Completado').count()
        # avg_rating = trips.filter(calificacion__isnull=False).aggregate(Avg('calificacion'))['calificacion__avg']
        
        statistics = {
            'conductor': {
                'id': driver.id,
                'nombre_completo': driver.get_nombre_completo(),
                'numero_documento': driver.numero_documento,
                'email': driver.user.email,
                'estado': driver.estado,
                'fecha_registro': driver.fecha_registro.strftime('%Y-%m-%d')
            },
            'vehiculo': {
                'placa': driver.vehiculo.placa if hasattr(driver, 'vehiculo') and driver.vehiculo else None,
                'marca': driver.vehiculo.marca if hasattr(driver, 'vehiculo') and driver.vehiculo else None,
                'modelo': driver.vehiculo.modelo if hasattr(driver, 'vehiculo') and driver.vehiculo else None,
            } if hasattr(driver, 'vehiculo') and driver.vehiculo else None,
            'stats': {
                'total_viajes': 247,  # Placeholder
                'calificacion_promedio': 4.8,
                'tasa_completado': 96,
                'total_reportes': 2,
                'viajes_mes_actual': 28,
                'ingresos_totales': 8950000
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': statistics
        }, status=200)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener estadísticas: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def driver_autocomplete_api(request):
    """
    API para autocompletado de conductores.
    Devuelve lista de conductores que coinciden con el término de búsqueda.
    """
    try:
        search_term = request.GET.get('q', '').strip()
        
        if not search_term or len(search_term) < 2:
            return JsonResponse({'results': []})
        
        # Buscar conductores que coincidan con el término
        conductores = Conductor.objects.select_related('user', 'vehiculo').filter(
            Q(user__first_name__icontains=search_term) |
            Q(user__last_name__icontains=search_term) |
            Q(segundo_nombre__icontains=search_term) |
            Q(segundo_apellido__icontains=search_term) |
            Q(numero_documento__icontains=search_term) |
            Q(vehiculo__placa__icontains=search_term)
        )[:10]  # Limitar a 10 resultados
        
        results = []
        for conductor in conductores:
            placa = conductor.vehiculo.placa if hasattr(conductor, 'vehiculo') and conductor.vehiculo else 'Sin placa'
            results.append({
                'id': conductor.id,
                'text': f"{conductor.get_nombre_completo()} - {conductor.numero_documento} - {placa}",
                'nombre_completo': conductor.get_nombre_completo(),
                'numero_documento': conductor.numero_documento,
                'placa': placa,
                'estado': conductor.estado
            })
        
        return JsonResponse({'results': results})
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error en búsqueda: {str(e)}'
        }, status=500)
        
# TODO: Obtener viajes reales
# trips = trips_queryset.select_related('cliente').order_by('-fecha')[:20]
from datetime import datetime, date, timedelta
import csv
import random

# Data mockeada de viajes realista
trips = []
base_date = datetime.now()

for i in range(5):
    # Fechas aleatorias en los últimos 30 días
    days_ago = random.randint(0, 30)
    trip_date = base_date - timedelta(days=days_ago, hours=random.randint(6, 22))
    
    # Estados posibles con probabilidades realistas
    status_weights = [('Completado', 0.85), ('Cancelado', 0.10), ('En Progreso', 0.04), ('Aceptado', 0.01)]
    status = random.choices([s[0] for s in status_weights], weights=[s[1] for s in status_weights])[0]
    
    # Calificación solo para viajes completados
    rating = random.choice([4.5, 5.0, 4.0, 4.5, 3.5, 5.0, 4.0, 4.5, 5.0, 3.0]) if status == 'Completado' and random.random() > 0.1 else None
    
    # Ubicaciones realistas en Bogotá
    locations = [
        'Centro Internacional', 'Aeropuerto El Dorado', 'Zona Rosa', 'Parque de la 93',
        'Centro Andino', 'Gran Estación', 'Plaza de las Américas', 'Calle 80',
        'Centro Empresarial Santa Fe', 'Parque Nacional', 'Usaquén', 'Chapinero',
        'Salitre Plaza', 'Portal Norte', 'Calle 170', 'Centro Histórico'
    ]
    
    origin = random.choice(locations)
    destination = random.choice([loc for loc in locations if loc != origin])
    
    # Clientes mockeados
    clients = [
        'María González', 'Carlos Rodríguez', 'Ana López', 'Juan Martínez', 
        'Laura Sánchez', 'Diego Pérez', 'Camila Torres', 'Andrés Ramírez',
        'Valentina Morales', 'Santiago Herrera', 'Isabella Castro', 'Mateo Vargas'
    ]
    
    trip = {
        'id': f'TRP{1000 + i}',
        'date': trip_date,
        'status': status,
        'rating': rating,
        'origin': origin,
        'destination': destination,
        'client': random.choice(clients),
        'amount': random.randint(15000, 45000),
        'distance': round(random.uniform(5.0, 25.0), 1),
        'duration': f"{random.randint(20, 90):02d}:{random.randint(0, 59):02d}:00"
    }
    trips.append(trip)
